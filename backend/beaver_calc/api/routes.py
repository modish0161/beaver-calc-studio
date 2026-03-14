"""
API routes for BeaverCalc Studio
"""
import uuid
from datetime import datetime
from typing import Dict, Any, Optional
import io

from flask import Blueprint, request, jsonify, send_file, current_app
from flask_jwt_extended import jwt_required, get_jwt_identity

from ..extensions import db
from ..models import Project, Calculator, Run, User, AuditLog, Template, SignOff
from ..calculators import registry
from ..reports import generate_pdf_report, generate_docx_report

api_bp = Blueprint('api', __name__)


# Projects endpoints
@api_bp.route('/projects', methods=['GET'])
@jwt_required()
def list_projects():
    """List all projects accessible to the current user"""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)

    # Role-based access: admins see all, others see only their own projects
    if user.role == 'admin':
        projects = Project.query.all()
    else:
        projects = Project.query.filter_by(created_by_id=user_id).all()

    return jsonify({
        'projects': [{
            'id': project.id,
            'name': project.name,
            'description': project.description,
            'client': project.client,
            'project_number': project.project_number,
            'status': project.status,
            'created_by': project.created_by.full_name if project.created_by else None,
            'created_at': project.created_at.isoformat()
        } for project in projects]
    }), 200


@api_bp.route('/projects', methods=['POST'])
@jwt_required()
def create_project():
    """Create a new project"""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)

    data = request.get_json()

    if not data or not data.get('name'):
        return jsonify({'error': 'Project name is required'}), 400

    project = Project(
        name=data['name'],
        description=data.get('description'),
        client=data.get('client'),
        project_number=data.get('project_number'),
        created_by_id=user_id
    )

    db.session.add(project)
    db.session.commit()

    # Log project creation
    audit_log = AuditLog(
        actor_id=user_id,
        action='create_project',
        resource_type='project',
        resource_id=project.id,
        details={'name': project.name},
        ip_address=request.remote_addr,
        user_agent=request.headers.get('User-Agent')
    )
    db.session.add(audit_log)
    db.session.commit()

    return jsonify({
        'message': 'Project created successfully',
        'project': {
            'id': project.id,
            'name': project.name,
            'description': project.description,
            'client': project.client,
            'project_number': project.project_number,
            'status': project.status,
            'created_at': project.created_at.isoformat()
        }
    }), 201


@api_bp.route('/projects/<int:project_id>', methods=['GET'])
@jwt_required()
def get_project(project_id):
    """Get a single project by ID"""
    project = Project.query.get(project_id)
    if not project:
        return jsonify({'error': 'Project not found'}), 404

    run_count = Run.query.filter_by(project_id=project.id).count()
    file_count = len(project.files)

    return jsonify({
        'project': {
            'id': project.id,
            'name': project.name,
            'description': project.description,
            'client': project.client,
            'project_number': project.project_number,
            'status': project.status,
            'created_by': project.created_by.full_name if project.created_by else None,
            'created_at': project.created_at.isoformat(),
            'updated_at': project.updated_at.isoformat(),
            'run_count': run_count,
            'file_count': file_count,
        }
    }), 200


@api_bp.route('/projects/<int:project_id>', methods=['PUT'])
@jwt_required()
def update_project(project_id):
    """Update a project"""
    user_id = int(get_jwt_identity())
    project = Project.query.get(project_id)
    if not project:
        return jsonify({'error': 'Project not found'}), 404

    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided'}), 400

    allowed_fields = ['name', 'description', 'client', 'project_number', 'status']
    for field in allowed_fields:
        if field in data:
            setattr(project, field, data[field])

    db.session.commit()

    audit_log = AuditLog(
        actor_id=user_id,
        action='update_project',
        resource_type='project',
        resource_id=project.id,
        details={k: data[k] for k in allowed_fields if k in data},
        ip_address=request.remote_addr,
        user_agent=request.headers.get('User-Agent')
    )
    db.session.add(audit_log)
    db.session.commit()

    return jsonify({
        'message': 'Project updated',
        'project': {
            'id': project.id,
            'name': project.name,
            'description': project.description,
            'client': project.client,
            'project_number': project.project_number,
            'status': project.status,
            'created_at': project.created_at.isoformat()
        }
    }), 200


@api_bp.route('/projects/<int:project_id>', methods=['DELETE'])
@jwt_required()
def delete_project(project_id):
    """Delete a project (admin or creator only)"""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    project = Project.query.get(project_id)
    if not project:
        return jsonify({'error': 'Project not found'}), 404

    if user.role != 'admin' and project.created_by_id != user_id:
        return jsonify({'error': 'Only admins or the project creator can delete'}), 403

    audit_log = AuditLog(
        actor_id=user_id,
        action='delete_project',
        resource_type='project',
        resource_id=project.id,
        details={'name': project.name},
        ip_address=request.remote_addr,
        user_agent=request.headers.get('User-Agent')
    )
    db.session.add(audit_log)

    db.session.delete(project)
    db.session.commit()

    return jsonify({'message': 'Project deleted'}), 200


# ── Materials & Sections API ────────────────────────────────────────────
# Static data served from memory — no database needed.

_STEEL_GRADES = {
    "S235": {"f_y": 235, "f_u": 360, "E": 210000},
    "S275": {"f_y": 275, "f_u": 410, "E": 210000},
    "S355": {"f_y": 355, "f_u": 470, "E": 210000},
    "S460": {"f_y": 460, "f_u": 540, "E": 210000},
}

_CONCRETE_GRADES = {
    "C20/25": {"f_ck": 20, "f_ctm": 2.2, "E_cm": 30000},
    "C25/30": {"f_ck": 25, "f_ctm": 2.6, "E_cm": 31000},
    "C28/35": {"f_ck": 28, "f_ctm": 2.8, "E_cm": 32000},
    "C30/37": {"f_ck": 30, "f_ctm": 2.9, "E_cm": 33000},
    "C32/40": {"f_ck": 32, "f_ctm": 3.0, "E_cm": 33500},
    "C35/45": {"f_ck": 35, "f_ctm": 3.2, "E_cm": 34000},
    "C40/50": {"f_ck": 40, "f_ctm": 3.5, "E_cm": 35000},
    "C50/60": {"f_ck": 50, "f_ctm": 4.1, "E_cm": 37000},
}

_TIMBER_GRADES = {
    "C16": {"f_m_k": 16, "f_c_0_k": 17, "E_0_mean": 8000},
    "C24": {"f_m_k": 24, "f_c_0_k": 21, "E_0_mean": 11000},
    "D30": {"f_m_k": 30, "f_c_0_k": 23, "E_0_mean": 10000},
    "GL24h": {"f_m_k": 24, "f_c_0_k": 24, "E_0_mean": 11500},
    "GL28h": {"f_m_k": 28, "f_c_0_k": 28, "E_0_mean": 12600},
}


@api_bp.route('/materials/steel', methods=['GET'])
def list_steel_grades():
    return jsonify({'grades': _STEEL_GRADES}), 200


@api_bp.route('/materials/concrete', methods=['GET'])
def list_concrete_grades():
    return jsonify({'grades': _CONCRETE_GRADES}), 200


@api_bp.route('/materials/timber', methods=['GET'])
def list_timber_grades():
    return jsonify({'grades': _TIMBER_GRADES}), 200


# Minimal UKB/UKC/PFC section catalogue (most common sizes).
# A production deployment would load the full Blue Book from a DB.

_UKB_SECTIONS = {
    "UKB 203x133x25": {"mass_kg_m": 25.1, "D": 203.2, "B": 133.2, "t_w": 5.7, "t_f": 7.8, "I_y": 2340, "I_z": 308, "W_el_y": 230},
    "UKB 254x146x31": {"mass_kg_m": 31.1, "D": 251.4, "B": 146.1, "t_w": 6.0, "t_f": 8.6, "I_y": 4410, "I_z": 448, "W_el_y": 351},
    "UKB 305x165x40": {"mass_kg_m": 40.3, "D": 303.4, "B": 165.0, "t_w": 6.0, "t_f": 10.2, "I_y": 8500, "I_z": 764, "W_el_y": 560},
    "UKB 356x171x51": {"mass_kg_m": 51.0, "D": 355.0, "B": 171.5, "t_w": 7.4, "t_f": 11.5, "I_y": 14100, "I_z": 968, "W_el_y": 796},
    "UKB 406x178x60": {"mass_kg_m": 60.1, "D": 406.4, "B": 177.9, "t_w": 7.9, "t_f": 12.8, "I_y": 21600, "I_z": 1200, "W_el_y": 1060},
    "UKB 457x191x67": {"mass_kg_m": 67.1, "D": 453.4, "B": 189.9, "t_w": 8.5, "t_f": 12.7, "I_y": 29400, "I_z": 1450, "W_el_y": 1300},
    "UKB 457x191x82": {"mass_kg_m": 82.0, "D": 460.0, "B": 191.3, "t_w": 9.9, "t_f": 16.0, "I_y": 37100, "I_z": 1870, "W_el_y": 1610},
    "UKB 533x210x82": {"mass_kg_m": 82.2, "D": 528.3, "B": 208.8, "t_w": 9.6, "t_f": 13.2, "I_y": 47500, "I_z": 2010, "W_el_y": 1800},
    "UKB 610x229x101": {"mass_kg_m": 101.2, "D": 602.6, "B": 227.6, "t_w": 10.5, "t_f": 14.8, "I_y": 75800, "I_z": 2910, "W_el_y": 2520},
}

_UKC_SECTIONS = {
    "UKC 152x152x23": {"mass_kg_m": 23.0, "D": 152.4, "B": 152.2, "t_w": 5.8, "t_f": 6.8, "I_y": 1250, "I_z": 429},
    "UKC 152x152x30": {"mass_kg_m": 30.0, "D": 157.6, "B": 152.9, "t_w": 6.5, "t_f": 9.4, "I_y": 1750, "I_z": 604},
    "UKC 203x203x46": {"mass_kg_m": 46.1, "D": 203.2, "B": 203.6, "t_w": 7.2, "t_f": 11.0, "I_y": 4570, "I_z": 1550},
    "UKC 203x203x60": {"mass_kg_m": 60.0, "D": 209.6, "B": 205.8, "t_w": 9.4, "t_f": 14.2, "I_y": 6130, "I_z": 2070},
    "UKC 254x254x73": {"mass_kg_m": 73.1, "D": 254.1, "B": 254.6, "t_w": 8.6, "t_f": 14.2, "I_y": 11400, "I_z": 3910},
    "UKC 305x305x97": {"mass_kg_m": 97.0, "D": 307.9, "B": 305.3, "t_w": 9.9, "t_f": 15.4, "I_y": 22300, "I_z": 7310},
}

_PFC_SECTIONS = {
    "PFC 150x75": {"mass_kg_m": 17.9, "D": 150, "B": 75, "t_w": 5.5, "t_f": 10.0, "I_y": 861, "I_z": 63.5},
    "PFC 200x75": {"mass_kg_m": 23.4, "D": 200, "B": 75, "t_w": 6.0, "t_f": 11.5, "I_y": 1520, "I_z": 71.8},
    "PFC 230x75": {"mass_kg_m": 26.1, "D": 230, "B": 75, "t_w": 6.5, "t_f": 12.5, "I_y": 2280, "I_z": 77.3},
    "PFC 260x75": {"mass_kg_m": 28.0, "D": 260, "B": 75, "t_w": 7.0, "t_f": 12.0, "I_y": 3120, "I_z": 72.4},
    "PFC 300x90": {"mass_kg_m": 41.4, "D": 300, "B": 90, "t_w": 9.0, "t_f": 15.5, "I_y": 7220, "I_z": 162},
}


@api_bp.route('/sections/ukb', methods=['GET'])
def list_ukb_sections():
    return jsonify({'sections': _UKB_SECTIONS}), 200


@api_bp.route('/sections/ukc', methods=['GET'])
def list_ukc_sections():
    return jsonify({'sections': _UKC_SECTIONS}), 200


@api_bp.route('/sections/pfc', methods=['GET'])
def list_pfc_sections():
    return jsonify({'sections': _PFC_SECTIONS}), 200


# ── Templates API ───────────────────────────────────────────────────────

@api_bp.route('/templates', methods=['GET'])
@jwt_required()
def list_templates():
    """List saved input templates"""
    calculator_key = request.args.get('calculator')
    query = Template.query.order_by(Template.updated_at.desc())
    if calculator_key:
        query = query.filter_by(calculator_key=calculator_key)
    templates = query.all()
    return jsonify({
        'templates': [{
            'id': t.id,
            'name': t.name,
            'calculator_key': t.calculator_key,
            'description': t.description,
            'inputs': t.inputs,
            'use_count': t.use_count,
            'created_by': t.created_by.full_name if t.created_by else None,
            'created_at': t.created_at.isoformat(),
        } for t in templates]
    }), 200


@api_bp.route('/templates', methods=['POST'])
@jwt_required()
def create_template():
    """Save current calculator inputs as a reusable template"""
    user_id = int(get_jwt_identity())
    data = request.get_json()

    if not data or not data.get('name') or not data.get('calculator_key') or not data.get('inputs'):
        return jsonify({'error': 'name, calculator_key, and inputs are required'}), 400

    template = Template(
        name=data['name'],
        calculator_key=data['calculator_key'],
        description=data.get('description'),
        inputs=data['inputs'],
        created_by_id=user_id,
        project_id=data.get('project_id'),
    )
    db.session.add(template)
    db.session.commit()

    return jsonify({
        'message': 'Template saved',
        'template': {
            'id': template.id,
            'name': template.name,
            'calculator_key': template.calculator_key,
        }
    }), 201


@api_bp.route('/templates/<int:template_id>', methods=['DELETE'])
@jwt_required()
def delete_template(template_id):
    """Delete a template"""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    template = Template.query.get(template_id)
    if not template:
        return jsonify({'error': 'Template not found'}), 404
    if user.role != 'admin' and template.created_by_id != user_id:
        return jsonify({'error': 'Only admins or the creator can delete'}), 403
    db.session.delete(template)
    db.session.commit()
    return jsonify({'message': 'Template deleted'}), 200


@api_bp.route('/templates/<int:template_id>/apply', methods=['POST'])
@jwt_required()
def apply_template(template_id):
    """Increment use_count and return template inputs"""
    template = Template.query.get(template_id)
    if not template:
        return jsonify({'error': 'Template not found'}), 404
    template.use_count += 1
    db.session.commit()
    return jsonify({
        'calculator_key': template.calculator_key,
        'inputs': template.inputs,
    }), 200


# Calculators endpoints
@api_bp.route('/calculators', methods=['GET'])
@jwt_required()
def list_calculators():
    """List all available calculators"""
    calculators = []
    for calc_plugin in registry.list_calculators().values():
        calculators.append(calc_plugin.get_metadata())

    return jsonify({
        'calculators': calculators
    }), 200


@api_bp.route('/calculators/<key>', methods=['GET'])
@jwt_required()
def get_calculator(key):
    """Get calculator metadata and schema"""
    calculator = registry.get_calculator(key)

    if not calculator:
        return jsonify({'error': 'Calculator not found'}), 404

    return jsonify({
        'calculator': calculator.get_metadata()
    }), 200


# Runs endpoints
@api_bp.route('/runs', methods=['POST'])
@jwt_required()
def create_run():
    """Start a new calculation run"""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)

    data = request.get_json()

    required_fields = ['calculator', 'project_id', 'inputs']
    if not all(field in data for field in required_fields):
        return jsonify({'error': 'calculator, project_id, and inputs are required'}), 400

    # Validate project exists
    project = Project.query.get(data['project_id'])
    if not project:
        return jsonify({'error': 'Project not found'}), 404

    # Validate calculator exists in plugin registry
    calc_plugin = registry.get_calculator(data['calculator'])
    if not calc_plugin:
        return jsonify({'error': 'Calculator not found'}), 404

    # Also check if calculator exists in database (for metadata)
    calculator = Calculator.query.filter_by(key=data['calculator'], is_active=True).first()
    calculator_id = calculator.id if calculator else None

    # Generate unique run ID
    run_id = str(uuid.uuid4())

    # Create run record
    run = Run(
        run_id=run_id,
        project_id=data['project_id'],
        calculator_id=calculator_id,
        user_id=user_id,
        inputs=data['inputs'],
        metadata_=data.get('metadata', {}),
        status='running'
    )
    run.started_at = datetime.utcnow()

    db.session.add(run)
    db.session.commit()

    try:
        # Emit progress via WebSocket (if SocketIO available)
        sio = current_app.extensions.get('socketio')
        if sio:
            from ..events import emit_run_progress
            emit_run_progress(sio, run_id, 10, 'Calculation started')

        # Perform the calculation
        results = calc_plugin.calculate(data['inputs'])

        if sio:
            emit_run_progress(sio, run_id, 90, 'Calculation complete, saving results')

        # Update run with results
        run.results = results
        run.status = 'completed'
        run.completed_at = datetime.utcnow()
        run.run_hash = Run.compute_hash(data['inputs'], results)

        db.session.commit()

        if sio:
            from ..events import emit_run_completed
            # Build a small summary for the WS payload
            summary_keys = [k for k in results if isinstance(results.get(k), (int, float, str, bool))][:8]
            emit_run_completed(sio, run_id, {k: results[k] for k in summary_keys})

    except Exception as e:
        # Handle calculation errors
        run.status = 'failed'
        run.error_message = str(e)
        run.completed_at = datetime.utcnow()
        db.session.commit()

        if sio:
            from ..events import emit_run_failed
            emit_run_failed(sio, run_id, str(e))

        return jsonify({
            'error': 'Calculation failed',
            'message': str(e),
            'run_id': run_id
        }), 500

    # Log run creation
    audit_log = AuditLog(
        actor_id=user_id,
        action='create_run',
        resource_type='run',
        resource_id=run.id,
        details={'calculator': data['calculator'], 'project_id': data['project_id']},
        ip_address=request.remote_addr,
        user_agent=request.headers.get('User-Agent')
    )
    db.session.add(audit_log)
    db.session.commit()

    return jsonify({
        'message': 'Run created successfully',
        'run': {
            'id': run.id,
            'run_id': run.run_id,
            'status': run.status,
            'created_at': run.created_at.isoformat()
        }
    }), 201


@api_bp.route('/runs/<run_id>', methods=['GET'])
@jwt_required()
def get_run(run_id):
    """Get run status and results"""
    user_id = int(get_jwt_identity())

    run = Run.query.filter_by(run_id=run_id).first()

    if not run:
        return jsonify({'error': 'Run not found'}), 404

    # Basic access control - user must be the creator or have appropriate role
    if run.user_id != user_id:
        user = User.query.get(user_id)
        if user.role not in ['admin', 'checker']:
            return jsonify({'error': 'Access denied'}), 403

    return jsonify({
        'run': {
            'id': run.id,
            'run_id': run.run_id,
            'project_id': run.project_id,
            'calculator_key': run.calculator.key if run.calculator else None,
            'status': run.status,
            'inputs': run.inputs,
            'results': run.results,
            'error_message': run.error_message,
            'metadata': run.metadata_,
            'run_hash': run.run_hash,
            'created_at': run.created_at.isoformat(),
            'started_at': run.started_at.isoformat() if run.started_at else None,
            'completed_at': run.completed_at.isoformat() if run.completed_at else None
        }
    }), 200


@api_bp.route('/runs/<run_id>/verify', methods=['GET'])
@jwt_required()
def verify_run(run_id):
    """Verify run integrity by recomputing the hash."""
    run = Run.query.filter_by(run_id=run_id).first()
    if not run:
        return jsonify({'error': 'Run not found'}), 404

    if not run.run_hash:
        return jsonify({'verified': False, 'reason': 'No hash stored for this run'}), 200

    expected = Run.compute_hash(run.inputs, run.results)
    verified = expected == run.run_hash

    return jsonify({
        'verified': verified,
        'run_id': run.run_id,
        'stored_hash': run.run_hash,
        'computed_hash': expected,
    }), 200


@api_bp.route('/runs', methods=['GET'])
@jwt_required()
def list_runs():
    """List runs for current user"""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)

    # Get query parameters
    project_id = request.args.get('project_id', type=int)
    calculator_key = request.args.get('calculator')
    status = request.args.get('status')
    limit = request.args.get('limit', 50, type=int)
    offset = request.args.get('offset', 0, type=int)

    query = Run.query

    # Filter by user (unless admin)
    if user.role != 'admin':
        query = query.filter_by(user_id=user_id)

    # Filter by project if specified
    if project_id:
        query = query.filter_by(project_id=project_id)

    # Filter by calculator key if specified
    if calculator_key:
        calc = Calculator.query.filter_by(key=calculator_key).first()
        if calc:
            query = query.filter_by(calculator_id=calc.id)

    # Filter by status if specified
    if status:
        query = query.filter_by(status=status)

    # Order by creation date, most recent first
    query = query.order_by(Run.created_at.desc())

    # Pagination
    total = query.count()
    runs = query.limit(limit).offset(offset).all()

    return jsonify({
        'runs': [{
            'id': run.id,
            'run_id': run.run_id,
            'project_id': run.project_id,
            'project_name': run.project.name if run.project else None,
            'calculator_key': run.calculator.key if run.calculator else None,
            'calculator_name': run.calculator.name if run.calculator else None,
            'status': run.status,
            'inputs': run.inputs,
            'results': run.results,
            'created_at': run.created_at.isoformat(),
            'completed_at': run.completed_at.isoformat() if run.completed_at else None
        } for run in runs],
        'total': total,
        'limit': limit,
        'offset': offset
    }), 200


@api_bp.route('/runs/sync', methods=['POST'])
@jwt_required()
def sync_run():
    """Store a pre-computed run (from frontend local calculation) for audit trail.

    Accepts inputs + results directly; does NOT re-run the calculator.
    project_id is optional — omit it for quick ad-hoc calculations.
    """
    user_id = int(get_jwt_identity())

    data = request.get_json()
    if not data or not data.get('calculator') or not data.get('inputs') or not data.get('results'):
        return jsonify({'error': 'calculator, inputs, and results are required'}), 400

    # Validate calculator exists
    calc_plugin = registry.get_calculator(data['calculator'])
    if not calc_plugin:
        return jsonify({'error': 'Calculator not found'}), 404

    calculator = Calculator.query.filter_by(key=data['calculator'], is_active=True).first()
    calculator_id = calculator.id if calculator else None

    # Optional project
    project_id = data.get('project_id')
    if project_id:
        project = Project.query.get(project_id)
        if not project:
            return jsonify({'error': 'Project not found'}), 404

    run_id = str(uuid.uuid4())
    now = datetime.utcnow()
    run = Run(
        run_id=run_id,
        project_id=project_id,
        calculator_id=calculator_id,
        user_id=user_id,
        inputs=data['inputs'],
        results=data['results'],
        metadata_=data.get('metadata', {}),
        status='completed',
        started_at=now,
        completed_at=now,
        run_hash=Run.compute_hash(data['inputs'], data['results']),
    )
    db.session.add(run)
    db.session.commit()

    # Audit
    audit_log = AuditLog(
        actor_id=user_id,
        action='sync_run',
        resource_type='run',
        resource_id=run.id,
        details={'calculator': data['calculator'], 'project_id': project_id},
        ip_address=request.remote_addr,
        user_agent=request.headers.get('User-Agent'),
    )
    db.session.add(audit_log)
    db.session.commit()

    return jsonify({
        'message': 'Run synced successfully',
        'run': {
            'id': run.id,
            'run_id': run.run_id,
            'status': run.status,
            'created_at': run.created_at.isoformat(),
        }
    }), 201


# Report generation endpoints
@api_bp.route('/runs/<string:run_id>/report/pdf', methods=['GET'])
@jwt_required()
def generate_pdf(run_id: str):
    """Generate PDF report for a calculation run"""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    
    # Get the run
    run = Run.query.filter_by(run_id=run_id).first()
    if not run:
        return jsonify({'error': 'Run not found'}), 404
    
    # Check access permissions
    if user.role != 'admin' and run.user_id != user_id:
        return jsonify({'error': 'Unauthorized'}), 403
    
    # Prepare run data for report
    run_data = _prepare_run_data_for_report(run, user)
    
    try:
        # Generate PDF
        pdf_bytes = generate_pdf_report(run_data)
        
        # Create audit log
        audit_log = AuditLog(
            user_id=user_id,
            action='generate_pdf_report',
            resource_type='run',
            resource_id=run.id,
            metadata_={'run_id': run_id, 'format': 'pdf'}
        )
        db.session.add(audit_log)
        db.session.commit()
        
        # Send PDF file
        filename = f"{run.calculator.key}_{run_id[:8]}_report.pdf"
        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({'error': f'Failed to generate PDF: {str(e)}'}), 500


@api_bp.route('/runs/<string:run_id>/report/docx', methods=['GET'])
@jwt_required()
def generate_docx(run_id: str):
    """Generate DOCX report for a calculation run"""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    
    # Get the run
    run = Run.query.filter_by(run_id=run_id).first()
    if not run:
        return jsonify({'error': 'Run not found'}), 404
    
    # Check access permissions
    if user.role != 'admin' and run.user_id != user_id:
        return jsonify({'error': 'Unauthorized'}), 403
    
    # Prepare run data for report
    run_data = _prepare_run_data_for_report(run, user)
    
    try:
        # Generate DOCX
        docx_bytes = generate_docx_report(run_data)
        
        # Create audit log
        audit_log = AuditLog(
            user_id=user_id,
            action='generate_docx_report',
            resource_type='run',
            resource_id=run.id,
            metadata_={'run_id': run_id, 'format': 'docx'}
        )
        db.session.add(audit_log)
        db.session.commit()
        
        # Send DOCX file
        filename = f"{run.calculator.key}_{run_id[:8]}_report.docx"
        return send_file(
            io.BytesIO(docx_bytes),
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({'error': f'Failed to generate DOCX: {str(e)}'}), 500


def _prepare_run_data_for_report(run: Run, user: User) -> Dict[str, Any]:
    """Prepare run data for report generation"""
    
    # Get calculator metadata
    calculator = run.calculator
    calculator_instance = registry.get_calculator(calculator.key)
    
    # Extract input parameters
    input_parameters = []
    if run.inputs and isinstance(run.inputs, dict):
        for key, value in run.inputs.items():
            # Try to get parameter metadata from calculator schema
            param_meta = {}
            if calculator_instance and hasattr(calculator_instance, 'get_input_schema'):
                schema = calculator_instance.get_input_schema()
                if key in schema.get('properties', {}):
                    prop = schema['properties'][key]
                    param_meta = {
                        'name': prop.get('title', key),
                        'symbol': prop.get('symbol', ''),
                        'unit': prop.get('unit', ''),
                        'reference': prop.get('reference', ''),
                    }
            
            input_parameters.append({
                'name': param_meta.get('name', key.replace('_', ' ').title()),
                'symbol': param_meta.get('symbol', ''),
                'value': value,
                'unit': param_meta.get('unit', ''),
                'reference': param_meta.get('reference', ''),
            })
    
    # Extract results
    results = []
    design_checks = []
    if run.results and isinstance(run.results, dict):
        # Extract main results
        for key, value in run.results.items():
            if key not in ['checks', 'summary', 'steps']:
                if isinstance(value, dict) and 'value' in value:
                    results.append({
                        'name': value.get('name', key.replace('_', ' ').title()),
                        'value': value.get('value'),
                        'unit': value.get('unit', ''),
                    })
                else:
                    results.append({
                        'name': key.replace('_', ' ').title(),
                        'value': value,
                        'unit': '',
                    })
        
        # Extract design checks
        if 'checks' in run.results:
            for check in run.results['checks']:
                design_checks.append({
                    'name': check.get('name', ''),
                    'capacity': check.get('capacity', 0),
                    'demand': check.get('demand', 0),
                    'utilization': check.get('utilization', 0),
                    'unit': check.get('unit', ''),
                    'status': check.get('status', 'UNKNOWN'),
                })
    
    # Extract calculation steps
    calculation_steps = []
    if run.results and 'steps' in run.results:
        calculation_steps = run.results['steps']
    
    # Get design summary
    design_summary = ''
    if run.results and 'summary' in run.results:
        design_summary = run.results['summary']
    else:
        design_summary = 'Calculation completed successfully. All design checks passed.'
    
    # Compile report data
    report_data = {
        'id': run.run_id,
        'calculator_name': calculator.name,
        'calculator_description': calculator.description,
        'project_name': run.project.name if run.project else 'Unnamed Project',
        'project_id': run.project.project_number if run.project else 'N/A',
        'client_name': run.project.client if run.project else 'N/A',
        'engineer_name': user.full_name,
        'design_code': calculator.eurocodes[0] if calculator.eurocodes else 'EN 1993-1-1',
        'created_at': run.created_at.strftime('%d %B %Y %H:%M'),
        'input_parameters': input_parameters,
        'calculation_steps': calculation_steps,
        'results': results,
        'checks': design_checks,
        'summary': design_summary,
    }
    
    return report_data


# ── DXF export endpoint ─────────────────────────────────────────────────

@api_bp.route('/runs/<string:run_id>/report/dxf', methods=['GET'])
@jwt_required()
def generate_dxf(run_id: str):
    """Generate DXF drawing for a completed run (pad footing, hole pattern, etc.)"""
    import ezdxf
    import math as _math

    user_id = int(get_jwt_identity())
    run = Run.query.filter_by(run_id=run_id).first()
    if not run:
        return jsonify({'error': 'Run not found'}), 404
    if not run.results:
        return jsonify({'error': 'Run has no results'}), 400

    doc = ezdxf.new('R2010')
    msp = doc.modelspace()

    # Layer setup
    doc.layers.add('OUTLINE', color=7)   # white
    doc.layers.add('DIMS', color=3)      # green
    doc.layers.add('NOTES', color=5)     # blue
    doc.layers.add('HATCHING', color=8)  # grey

    calc_key = run.calculator.key if run.calculator else ''
    inputs = run.inputs or {}
    results = run.results or {}

    if 'pad_footing' in calc_key:
        # Pad footing plan view
        L = inputs.get('footing_length_m', 2.0)
        W = inputs.get('footing_width_m', 2.0)
        # Footing outline (metres)
        msp.add_lwpolyline(
            [(0, 0), (L, 0), (L, W), (0, W), (0, 0)],
            dxfattribs={'layer': 'OUTLINE'},
        )
        # Dimension lines
        msp.add_text(f'{L:.2f} m', dxfattribs={'layer': 'DIMS', 'height': 0.12}).set_placement((L / 2, -0.25))
        msp.add_text(f'{W:.2f} m', dxfattribs={'layer': 'DIMS', 'height': 0.12}).set_placement((-0.35, W / 2))
        # Centre cross
        msp.add_line((L / 2 - 0.15, W / 2), (L / 2 + 0.15, W / 2), dxfattribs={'layer': 'DIMS'})
        msp.add_line((L / 2, W / 2 - 0.15), (L / 2, W / 2 + 0.15), dxfattribs={'layer': 'DIMS'})
        # Title block
        sigma = results.get('sigma_v_kN_m2', 0)
        util = results.get('utilisation_bearing', 0)
        msp.add_text(
            f'Pad Footing {L}m x {W}m  |  σ={sigma:.1f} kPa  |  Util={util:.0%}',
            dxfattribs={'layer': 'NOTES', 'height': 0.10},
        ).set_placement((0, -0.6))
    elif 'hole_pattern' in calc_key:
        # Bolt hole pattern — plate outline + holes
        pw = inputs.get('plate_width_mm', 200)
        ph = inputs.get('plate_height_mm', 300)
        msp.add_lwpolyline(
            [(0, 0), (pw, 0), (pw, ph), (0, ph), (0, 0)],
            dxfattribs={'layer': 'OUTLINE'},
        )
        for hole in results.get('holes', []):
            msp.add_circle(
                (hole['x_mm'], hole['y_mm']),
                hole['dia_mm'] / 2,
                dxfattribs={'layer': 'OUTLINE'},
            )
    else:
        # Generic: dump key results as text
        y = 0
        for key, val in results.items():
            if isinstance(val, (int, float, str, bool)):
                msp.add_text(f'{key}: {val}', dxfattribs={'layer': 'NOTES', 'height': 2.5}).set_placement((0, y))
                y -= 5

    stream = io.StringIO()
    doc.write(stream)
    buf = io.BytesIO(stream.getvalue().encode(doc.output_encoding, errors='dxfreplace'))
    buf.seek(0)

    filename = f"{calc_key}_{run_id[:8]}.dxf"
    return send_file(buf, mimetype='application/dxf', as_attachment=True, download_name=filename)


# ── XLSX export endpoint ────────────────────────────────────────────────

@api_bp.route('/runs/<string:run_id>/report/xlsx', methods=['GET'])
@jwt_required()
def generate_xlsx(run_id: str):
    """Export run inputs + results as an Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    user_id = int(get_jwt_identity())
    run = Run.query.filter_by(run_id=run_id).first()
    if not run:
        return jsonify({'error': 'Run not found'}), 404

    wb = Workbook()

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')

    # ── Inputs sheet ────────────────────────────────────────────────────
    ws_in = wb.active
    ws_in.title = 'Inputs'
    ws_in.append(['Parameter', 'Value'])
    ws_in['A1'].font = header_font_white
    ws_in['B1'].font = header_font_white
    ws_in['A1'].fill = header_fill
    ws_in['B1'].fill = header_fill
    for k, v in (run.inputs or {}).items():
        ws_in.append([k.replace('_', ' ').title(), v])
    ws_in.column_dimensions['A'].width = 35
    ws_in.column_dimensions['B'].width = 20

    # ── Results sheet ───────────────────────────────────────────────────
    ws_res = wb.create_sheet('Results')
    ws_res.append(['Result', 'Value'])
    ws_res['A1'].font = header_font_white
    ws_res['B1'].font = header_font_white
    ws_res['A1'].fill = header_fill
    ws_res['B1'].fill = header_fill
    for k, v in (run.results or {}).items():
        if isinstance(v, (int, float, str, bool)):
            ws_res.append([k.replace('_', ' ').title(), v])
        elif isinstance(v, list):
            ws_res.append([k.replace('_', ' ').title(), str(v)])
    ws_res.column_dimensions['A'].width = 35
    ws_res.column_dimensions['B'].width = 30

    # ── Checks sheet (if present) ───────────────────────────────────────
    checks = (run.results or {}).get('checks', [])
    if checks and isinstance(checks, list):
        ws_chk = wb.create_sheet('Design Checks')
        ws_chk.append(['Check', 'Status', 'Utilisation', 'Detail'])
        for i, cell in enumerate(['A1', 'B1', 'C1', 'D1']):
            ws_chk[cell].font = header_font_white
            ws_chk[cell].fill = header_fill
        for chk in checks:
            if isinstance(chk, dict):
                ws_chk.append([
                    chk.get('name', ''),
                    chk.get('status', ''),
                    chk.get('utilisation', chk.get('utilization', '')),
                    chk.get('detail', ''),
                ])
        ws_chk.column_dimensions['A'].width = 35
        ws_chk.column_dimensions['B'].width = 12
        ws_chk.column_dimensions['C'].width = 14
        ws_chk.column_dimensions['D'].width = 45

    # ── Summary sheet ───────────────────────────────────────────────────
    ws_sum = wb.create_sheet('Summary')
    calc_name = run.calculator.name if run.calculator else 'Unknown'
    proj_name = run.project.name if run.project else 'N/A'
    ws_sum.append(['Field', 'Value'])
    ws_sum['A1'].font = header_font_white
    ws_sum['B1'].font = header_font_white
    ws_sum['A1'].fill = header_fill
    ws_sum['B1'].fill = header_fill
    ws_sum.append(['Calculator', calc_name])
    ws_sum.append(['Project', proj_name])
    ws_sum.append(['Run ID', run.run_id])
    ws_sum.append(['Status', run.status])
    ws_sum.append(['Created', run.created_at.strftime('%d %B %Y %H:%M')])
    ws_sum.append(['Run Hash', run.run_hash or 'N/A'])
    ws_sum.column_dimensions['A'].width = 20
    ws_sum.column_dimensions['B'].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    calc_key = run.calculator.key if run.calculator else 'run'
    filename = f"{calc_key}_{run_id[:8]}.xlsx"
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


# ── Sign-off / approval workflow ────────────────────────────────────────

@api_bp.route('/runs/<string:run_id>/signoffs', methods=['GET'])
@jwt_required()
def list_signoffs(run_id: str):
    """List all sign-offs for a run."""
    run = Run.query.filter_by(run_id=run_id).first()
    if not run:
        return jsonify({'error': 'Run not found'}), 404

    signoffs = SignOff.query.filter_by(run_id=run.id).all()
    return jsonify({
        'signoffs': [
            {
                'id': s.id,
                'role': s.role,
                'user': s.user.full_name,
                'user_id': s.user_id,
                'comment': s.comment,
                'signed_at': s.signed_at.isoformat(),
            }
            for s in signoffs
        ]
    }), 200


@api_bp.route('/runs/<string:run_id>/signoffs', methods=['POST'])
@jwt_required()
def create_signoff(run_id: str):
    """Sign off a run as designer, checker, or approver."""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)

    run = Run.query.filter_by(run_id=run_id).first()
    if not run:
        return jsonify({'error': 'Run not found'}), 404
    if run.status != 'completed':
        return jsonify({'error': 'Can only sign off completed runs'}), 400

    data = request.get_json() or {}
    role = data.get('role')
    if role not in ('designer', 'checker', 'approver'):
        return jsonify({'error': 'Role must be designer, checker, or approver'}), 400

    # Role-based permission: checkers+ can sign as checker; only admin/approver for approver
    allowed = {
        'designer': ['admin', 'designer', 'checker'],
        'checker': ['admin', 'checker'],
        'approver': ['admin'],
    }
    if user.role not in allowed.get(role, []):
        return jsonify({'error': f'Your role ({user.role}) cannot sign off as {role}'}), 403

    # Enforce ordering: designer before checker, checker before approver
    existing = {s.role for s in SignOff.query.filter_by(run_id=run.id).all()}
    if role == 'checker' and 'designer' not in existing:
        return jsonify({'error': 'Designer must sign off before checker'}), 400
    if role == 'approver' and 'checker' not in existing:
        return jsonify({'error': 'Checker must sign off before approver'}), 400

    # Prevent duplicate sign-off for same role
    if role in existing:
        return jsonify({'error': f'{role.title()} sign-off already exists'}), 409

    signoff = SignOff(
        run_id=run.id,
        role=role,
        user_id=user_id,
        comment=data.get('comment', ''),
    )
    db.session.add(signoff)

    # Audit
    audit_log = AuditLog(
        actor_id=user_id,
        action=f'signoff_{role}',
        resource_type='run',
        resource_id=run.id,
        details={'run_id': run_id, 'role': role},
        ip_address=request.remote_addr,
        user_agent=request.headers.get('User-Agent'),
    )
    db.session.add(audit_log)
    db.session.commit()

    return jsonify({
        'message': f'{role.title()} sign-off recorded',
        'signoff': {
            'id': signoff.id,
            'role': signoff.role,
            'user': user.full_name,
            'signed_at': signoff.signed_at.isoformat(),
        },
    }), 201


# ── Audit log query endpoint ────────────────────────────────────────────

@api_bp.route('/audit', methods=['GET'])
@jwt_required()
def list_audit_logs():
    """List audit log entries (admin and checker roles only)."""
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)

    if user.role not in ('admin', 'checker'):
        return jsonify({'error': 'Admin or checker access required'}), 403

    # Query params
    action = request.args.get('action')
    resource_type = request.args.get('resource_type')
    limit = request.args.get('limit', 100, type=int)
    offset = request.args.get('offset', 0, type=int)

    query = AuditLog.query.order_by(AuditLog.created_at.desc())
    if action:
        query = query.filter_by(action=action)
    if resource_type:
        query = query.filter_by(resource_type=resource_type)

    total = query.count()
    logs = query.limit(limit).offset(offset).all()

    return jsonify({
        'audit_logs': [{
            'id': log.id,
            'actor': log.actor.full_name if log.actor else 'System',
            'actor_id': log.actor_id,
            'action': log.action,
            'resource_type': log.resource_type,
            'resource_id': log.resource_id,
            'details': log.details,
            'ip_address': log.ip_address,
            'created_at': log.created_at.isoformat(),
        } for log in logs],
        'total': total,
        'limit': limit,
        'offset': offset,
    }), 200


# ── Project files upload/download ───────────────────────────────────────

import hashlib
import os

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'instance', 'uploads')

ALLOWED_EXTENSIONS = {'pdf', 'docx', 'xlsx', 'dxf', 'png', 'jpg', 'jpeg', 'dwg'}
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50 MB


def _allowed_file(filename: str) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@api_bp.route('/projects/<int:project_id>/files', methods=['GET'])
@jwt_required()
def list_project_files(project_id):
    """List all files attached to a project."""
    from ..models import File
    project = Project.query.get(project_id)
    if not project:
        return jsonify({'error': 'Project not found'}), 404

    files = File.query.filter_by(project_id=project_id).order_by(File.created_at.desc()).all()
    return jsonify({
        'files': [{
            'id': f.id,
            'filename': f.original_filename,
            'file_type': f.file_type,
            'file_size': f.file_size,
            'mime_type': f.mime_type,
            'uploaded_by': f.uploaded_by.full_name if f.uploaded_by else None,
            'created_at': f.created_at.isoformat(),
        } for f in files]
    }), 200


@api_bp.route('/projects/<int:project_id>/files', methods=['POST'])
@jwt_required()
def upload_project_file(project_id):
    """Upload a file (drawing, photo, document) to a project."""
    from ..models import File
    import uuid as _uuid

    user_id = int(get_jwt_identity())
    project = Project.query.get(project_id)
    if not project:
        return jsonify({'error': 'Project not found'}), 404

    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    if file.filename == '' or not file.filename:
        return jsonify({'error': 'No file selected'}), 400

    if not _allowed_file(file.filename):
        return jsonify({'error': f'File type not allowed. Allowed: {", ".join(sorted(ALLOWED_EXTENSIONS))}'}), 400

    # Read and validate size
    file_bytes = file.read()
    if len(file_bytes) > MAX_FILE_SIZE:
        return jsonify({'error': 'File exceeds 50 MB limit'}), 400

    # Compute hash
    sha256 = hashlib.sha256(file_bytes).hexdigest()

    # Determine file type
    ext = file.filename.rsplit('.', 1)[1].lower()
    type_map = {
        'pdf': 'pdf', 'docx': 'docx', 'xlsx': 'xlsx', 'dxf': 'dxf',
        'png': 'image', 'jpg': 'image', 'jpeg': 'image', 'dwg': 'other',
    }
    file_type = type_map.get(ext, 'other')

    # Save to disk
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    stored_name = f"{_uuid.uuid4().hex}_{file.filename}"
    file_path = os.path.join(UPLOAD_DIR, stored_name)
    with open(file_path, 'wb') as f:
        f.write(file_bytes)

    file_record = File(
        filename=stored_name,
        original_filename=file.filename,
        file_path=file_path,
        file_type=file_type,
        file_size=len(file_bytes),
        mime_type=file.content_type or 'application/octet-stream',
        hash_sha256=sha256,
        project_id=project_id,
        uploaded_by_id=user_id,
    )
    db.session.add(file_record)

    audit_log = AuditLog(
        actor_id=user_id,
        action='upload_file',
        resource_type='file',
        resource_id=project_id,
        details={'filename': file.filename, 'size': len(file_bytes)},
        ip_address=request.remote_addr,
        user_agent=request.headers.get('User-Agent'),
    )
    db.session.add(audit_log)
    db.session.commit()

    return jsonify({
        'message': 'File uploaded',
        'file': {
            'id': file_record.id,
            'filename': file_record.original_filename,
            'file_type': file_record.file_type,
            'file_size': file_record.file_size,
        }
    }), 201


@api_bp.route('/projects/<int:project_id>/files/<int:file_id>', methods=['GET'])
@jwt_required()
def download_project_file(project_id, file_id):
    """Download a specific project file."""
    from ..models import File
    file_record = File.query.filter_by(id=file_id, project_id=project_id).first()
    if not file_record:
        return jsonify({'error': 'File not found'}), 404

    if not os.path.exists(file_record.file_path):
        return jsonify({'error': 'File not found on disk'}), 404

    return send_file(
        file_record.file_path,
        mimetype=file_record.mime_type,
        as_attachment=True,
        download_name=file_record.original_filename,
    )
